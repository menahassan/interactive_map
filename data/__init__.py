#!/usr/bin/env python

import openpyxl
#from googletrans import Translator
book1 = openpyxl.load_workbook('Human_Trafficking.xlsx')
book2 = openpyxl.load_workbook('Population.xlsx')

sheet1 = book1.active
sheet2 = book2.active
#translated = {} #list of hello translated in different languages
trafficking = {} #keys: countries , values: long/lat dictionary - DICT has not yet been tested
countriesPop = {} #keys: countries , values: long/lat dictionary - DICT has not yet been tested

for row in sheet2.iter_rows(min_row=5, max_row = 200, min_col=1, max_col=44, values_only=True):
    countryPop = [row[0]]
    for pop in row[4:]:
            countryPop = countryPop + [pop]
    countriesPop[row[1]] = countryPop

for row in sheet1.iter_rows(min_row=2, max_row = 168, min_col=1, max_col=4, values_only=True):
    traffic = [row[0]]
    for num in row[2:]:
            traffic = traffic + [num]
    trafficking[row[1]] = traffic
#for i in a:
#    langs = {}
#    translator = Translator()
#    translation = translator.translate('hello', dest=str(i[17]).lower())
#    txt = translation.text
#    langs["word"] = txt
#    langs["language"] = str(i[17])
#    translated[i[13]] = langs
#arr[0] is ['Active', 'AF', "Cote d'Ivoire"]
#arr[0][0] is 'Active'