import openpyxl

book = openpyxl.load_workbook('Embassies Consulates and Missions Lat Longs.xlsx')
sheet = book.active
DICT = {}
embassies_consulates = [] #list of rows as shown in excel spreadsheet
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=15):
    col = []
    for cell in row:
        col.append(cell.value)
    embassies_consulates.append(col)

for row in sheet.iter_rows(min_row=2, max_row = 276, min_col=1, max_col=19, values_only=True):
    data = {}
    if(row[14] != None and row[13] != None and row[18] != None):
        data["long"] = row[14]
        data["lat"]= row[13]
        data["yearOpen"] = row[18].year
        data['city'] = row[3]
        
        DICT[row[2]] = data

book = openpyxl.load_workbook('nodiplpresence.xlsx')
sheet = book.active

nodiplpresencelist = []
for row in sheet.iter_rows(min_row=3, min_col=1, max_row=sheet.max_row, max_col=3):
    col = []
    for cell in row:
        col.append(cell.value)
    nodiplpresencelist.append(col)

book = openpyxl.load_workbook('airpollution.xlsx')
sheet = book.active

airpollution = []
for row in sheet.iter_rows(min_row=3, min_col=1, max_row=11, max_col=2):
    col = []
    for cell in row:
        col.append(cell.value)
    airpollution.append(col)

book = openpyxl.load_workbook('co2emissions.xlsx')
sheet = book.active

co2emissions = []
for row in sheet.iter_rows(min_row=3, min_col=1, max_row=11, max_col=2):
    col = []
    for cell in row:
        col.append(cell.value)
    co2emissions.append(col)

book = openpyxl.load_workbook('povertyrate.xlsx')
sheet = book.active

povertyrate = []
for row in sheet.iter_rows(min_row=3, min_col=1, max_row=11, max_col=2):
    col = []
    for cell in row:
        col.append(cell.value)
    povertyrate.append(col)

book = openpyxl.load_workbook('issues_summaries.xlsx')
sheet = book.active

issues_summaries = []
for row in sheet.iter_rows(min_row=3, min_col=1, max_row=5, max_col=4):
    col = []
    for cell in row:
        col.append(cell.value)
    issues_summaries.append(col)

#arr[0] is ['Active', 'AF', "Cote d'Ivoire"]
#arr[0][0] is 'Active'

#book.save('iterbycols.xlsx')
book2 = openpyxl.load_workbook('Human_Development_Index.xlsx')

sheet2 = book2.active
countriesHDI = {} #keys: countries initial , values: first element is the name of the country, the rest is HDI starting from 1990 - 2018

for row in sheet2.iter_rows(min_row=3, max_row = 197, min_col=1, max_col=32, values_only=True):
    countryHDI = [row[1]]
    for hdi in row[3:]:
            countryHDI = countryHDI + [hdi]
    countriesHDI[row[2]] = countryHDI
